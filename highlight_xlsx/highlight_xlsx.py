import os
import sys
import glob
import pandas as pd
import openpyxl # Need to read xlsx
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

SUFFIX = "_highlighted"  # 出力に付ける印 (入力として拾わないために使う)


COLORS = {
    "white" : 'FFFFFF',
    "purple": 'FF00FF',
    "yellow": 'FFFF00',
    "red"   : 'FF0000',
    "sky"   : '00FFFF',
    "blue"  : '0000FF',
    "green" : '00FF00',
    "gray"  : 'CCCCCC',
}
DEFAULT_COLOR = 'FFFF00'  # 知らない色名はこれ (黄)


def convert_color_hex(color):
    """
    Converts a color name into a 6 digit hex string.
    Args:
        color (str): The name of the color (e.g. "red"), or "#RRGGBB".
    Returns:
        str: The 6 digit hex string (e.g. 'FF0000').
    Example:
        >>> convert_color_hex("red")
        'FF0000'
        >>> convert_color_hex("#ff0000")
        'FF0000'
        >>> convert_color_hex("unknown")
        'FFFF00'
    """
    if isinstance(color, str) and color.startswith("#"):
        return color[1:].upper()
    try:
        return COLORS[color]
    except (KeyError, TypeError):
        return DEFAULT_COLOR


def out_path(path, suffix="", ext=None, out_dir=None):
    """Makes an output path from an input path.
    Args:
        path (str): The input file path (with or without directories).
        suffix (str): The string to be added to the file body (e.g. "_highlighted").
        ext (str, optional): The output extension (e.g. ".csv"). Keeps the input one if None.
        out_dir (str, optional): The output directory. Uses the input one if None.
    Returns:
        str: The output file path.
    Example:
        >>> out_path("a.pdf.d/01.pdf", "_highlighted")
        "a.pdf.d/01_highlighted.pdf"
        >>> out_path("x/mtcars.pdf", "_1_1", ".csv", "csv")
        "csv/mtcars_1_1.csv"
    """
    dir_name, base = os.path.split(path)
    body, org_ext = os.path.splitext(base)
    name = f"{body}{suffix}{ext or org_ext}"
    if out_dir is not None:
        return os.path.join(out_dir, name)
    return os.path.join(dir_name, name) if dir_name else name


def convert_color_name(color):
    """
    Converts a color name to the color code of openpyxl.
    Args:
        color (str): The name of the color (e.g. "red"), or "#RRGGBB".
    Returns:
        str: The color code (e.g., 'FF0000' for red).
    Example:
        >>> convert_color_name("red")
        'FF0000'
        >>> convert_color_name("unknown")
        'FFFF00'  # default to yellow
    """
    return convert_color_hex(color)


def read_excel(path):
    """
    Reads an Excel file.
    Args:
        path (str): The path to the Excel file.
    Returns:
        pandas.DataFrame: The loaded DataFrame.
        None: If the file is not found or an error occurs.
    Raises:
        FileNotFoundError: If the file is not found.
        Exception: If any other error occurs.
    """
    try:
        df = pd.read_excel(path)
        return df
    except FileNotFoundError as e:
        print(f"File Not Found Error: {e}")
        input("Press Any Key")
    except Exception as e:
        print(f"Error: {e}")
        input("Press Any Key")


def highlight_xlsx(path_xlsx, keywords, colors, opacity = 0.3):
    out_xlsx = out_path(path_xlsx, SUFFIX)
    wb = openpyxl.load_workbook(path_xlsx)
    sheets = wb.worksheets
    offset = 64 # need to convert number to character
    for sheet in sheets:
        max_row = sheet.max_row
        max_col = sheet.max_column
        range_str = "".join([chr(1 + offset), str(1), ":", chr(max_col + offset), str(max_row)])
        for kwd, clr in zip(keywords, colors):
            highlight_cell(sheet, range_str, str(kwd), convert_color_name(clr))
    wb.save(out_xlsx)
    return out_xlsx

def highlight_cell(sheet, range_str, keyword, color):
    color_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    condition = f'EXACT("{keyword}", A1)'
    rule = FormulaRule(formula=[condition], fill=color_fill)
    sheet.conditional_formatting.add(range_str, rule)

def input_files(pattern, path_setting=None):
    """Lists the input files, without the outputs of the previous runs.
    Args:
        pattern (str): The glob pattern (e.g. "*.pdf").
        path_setting (str, optional): The setting file to be excluded.
    Returns:
        list: The paths to be processed.
    """
    files = []
    for path in sorted(glob.glob(pattern)):
        if SUFFIX in os.path.splitext(os.path.basename(path))[0]:
            continue
        if path_setting is not None and os.path.abspath(path) == os.path.abspath(path_setting):
            continue
        files.append(path)
    return files


def main(path_xlsx="highlight_xlsx.xlsx"):
    """Reads the setting xlsx and highlights the xlsx files in the current directory.
    Args:
        path_xlsx (str): The path to the setting xlsx file.
    Returns:
        int: 0 on success, 1 if the setting cannot be used.
    """
    df = read_excel(path_xlsx)
    if df is None:
        return 1
    df = df.dropna(subset=["keywords", "colors"])
    if len(df) == 0:
        print(f"No keywords in {path_xlsx}")
        input("Press Any Key")
        return 1
    for xlsx in input_files("*.xlsx", path_xlsx):
        out_xlsx = highlight_xlsx(xlsx, df.keywords, df.colors)
        os.startfile(out_xlsx)
    return 0


if __name__ == "__main__":
    sys.exit(main())
