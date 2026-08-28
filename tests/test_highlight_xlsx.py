"""highlight_xlsx のテスト."""
import openpyxl
import pytest

from conftest import load

hx = load("highlight_xlsx/highlight_xlsx.py")


class TestConvertColorName:
    @pytest.mark.parametrize(
        "name", ["white", "purple", "yellow", "red", "sky", "blue", "green"]
    )
    def test_returns_hex(self, name):
        """色名を 6桁の16進へ変換する."""
        col = hx.convert_color_name(name)
        assert len(col) == 6
        int(col, 16)

    def test_unknown_color_falls_back_to_yellow(self):
        """未知の色名は黄色になる."""
        assert hx.convert_color_name("no_such_color") == "FFFF00"


def make_xlsx(path, n_col=3):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, n_col + 1):
        ws.cell(row=1, column=i, value=f"c{i}")
    ws.cell(row=2, column=1, value="atari")
    wb.save(path)
    return str(path)


class TestHighlightXlsx:
    @pytest.mark.xfail(
        strict=True,
        reason="load_workbook(xlsx) が引数ではなく大域の xlsx を見ているため，"
        "呼び出しただけでは動かない",
    )
    def test_can_be_called_as_a_function(self, tmp_path):
        """関数として呼べる (path_xlsx を使う)."""
        path = make_xlsx(tmp_path / "in.xlsx")
        out = hx.highlight_xlsx(path, ["atari"], ["red"])
        assert openpyxl.load_workbook(out) is not None

    @pytest.mark.xfail(
        strict=True,
        reason="範囲を chr(max_col + 64) で作るので，27列以上で列名が壊れる",
    )
    def test_wide_sheet(self, tmp_path, monkeypatch):
        """27列以上のシートでも範囲が壊れない."""
        path = make_xlsx(tmp_path / "wide.xlsx", n_col=30)
        monkeypatch.setattr(hx, "xlsx", path, raising=False)
        out = hx.highlight_xlsx(path, ["atari"], ["red"])
        wb = openpyxl.load_workbook(out)
        ranges = [str(r) for r in wb.active.conditional_formatting]
        assert all("[" not in r for r in ranges)
